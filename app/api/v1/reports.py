"""
app/api/v1/reports.py
Endpoints báo cáo, thống kê và xuất Excel.
"""

import os
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Optional

from app.core.config import settings
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.attendance import AttendanceLog
from app.services.attendance import (
    get_logs_by_date, get_summary_today,
    get_log_by_id, update_attendance_log,
    delete_attendance_log, create_manual_attendance_log,
)


router = APIRouter(prefix="/api", tags=["reports"])

_bearer_opt = HTTPBearer(auto_error=False)


def _optional_user(
    creds: HTTPAuthorizationCredentials = Depends(_bearer_opt),
    db: Session = Depends(get_db),
):
    """Dependency tuỳ chọn: trả về user nếu có token, None nếu không có."""
    if creds is None:
        return None
    try:
        from app.core.security import decode_access_token
        from app.models.user import User
        payload = decode_access_token(creds.credentials)
        return db.query(User).filter_by(id=int(payload["sub"]), is_active=True).first()
    except Exception:
        return None


@router.get("/attendance")
def get_attendance(date: str = None, emp_code: str = None, days: int = 1,
                   current_user=Depends(_optional_user)):
    if date:
        logs = get_logs_by_date(date, emp_code)
    else:
        logs = []
        for i in range(days):
            d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            logs.extend(get_logs_by_date(d, emp_code))
    return {"logs": logs, "total": len(logs)}


@router.get("/summary")
def summary_today(current_user=Depends(get_current_user)):
    return get_summary_today()


@router.get("/summary/range")
def summary_range(from_date: str, to_date: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    start = datetime.strptime(from_date, "%Y-%m-%d").replace(hour=0,  minute=0)
    end   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59)
    logs  = db.query(AttendanceLog).filter(
        AttendanceLog.timestamp >= start,
        AttendanceLog.timestamp <= end,
    ).all()

    by_date: dict[str, dict] = {}
    for log in logs:
        day = log.timestamp.strftime("%Y-%m-%d")
        if day not in by_date:
            by_date[day] = {"check_in": set(), "check_out": set()}
        by_date[day][log.check_type].add(log.emp_code)

    dept_stats: dict[str, int] = {}
    for log in logs:
        dept = log.department or "Chưa xác định"
        dept_stats[dept] = dept_stats.get(dept, 0) + 1

    return {
        "from_date":  from_date,
        "to_date":    to_date,
        "total_logs": len(logs),
        "by_date": [
            {"date": d, "checked_in": len(v["check_in"]), "checked_out": len(v["check_out"])}
            for d, v in sorted(by_date.items())
        ],
        "by_dept": [{"dept": k, "count": v} for k, v in dept_stats.items()],
    }


@router.get("/reports/export")
def export_excel(from_date: str, to_date: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"error": "Cài openpyxl: pip install openpyxl"}, status_code=500)

    start = datetime.strptime(from_date, "%Y-%m-%d").replace(hour=0)
    end   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59)
    logs  = db.query(AttendanceLog).filter(
        AttendanceLog.timestamp >= start,
        AttendanceLog.timestamp <= end,
    ).order_by(AttendanceLog.timestamp).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo chấm công"

    ws.merge_cells("A1:G1")
    ws["A1"]           = f"BÁO CÁO CHẤM CÔNG  —  {from_date} đến {to_date}"
    ws["A1"].font      = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    thin   = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1A365D")
    headers  = ["STT", "Mã NV", "Họ tên", "Phòng ban", "Loại", "Thời gian", "Trạng thái"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    for i, log in enumerate(logs, 1):
        fill_color = "F0FFF4" if log.check_type == "check_in" else "EBF4FF"
        row_fill   = PatternFill("solid", fgColor=fill_color)
        for col, val in enumerate([
            i, log.emp_code, log.emp_name, log.department,
            "Vào" if log.check_type == "check_in" else "Ra",
            log.timestamp.strftime("%H:%M:%S  %d/%m/%Y"),
            log.note or "",
        ], 1):
            cell = ws.cell(row=i + 2, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
            cell.fill      = row_fill

    for col, w in enumerate([6, 10, 22, 18, 8, 24, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    settings.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filepath = str(settings.EXPORTS_DIR / f"chamcong_{from_date}_{to_date}.xlsx")
    wb.save(filepath)

    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"BaoCaoChamCong_{from_date}_{to_date}.xlsx",
    )


# ── GET /api/reports/employee/{emp_code}/stats ───────────────────

@router.get("/reports/employee/{emp_code}/stats")
def employee_stats(
    emp_code: str,
    year: int = 0,
    month: int = 0,
    db: Session = Depends(get_db),
    current_user=Depends(_optional_user),
):
    from app.services.work_calendar import get_employee_stats, get_employee_stats_month
    from datetime import date as _date
    y = year  or _date.today().year
    m = month or 0

    if m:
        return get_employee_stats_month(emp_code, y, m, db)
    return get_employee_stats(emp_code, y, db)


# ── Schemas điểm danh thủ công ────────────────────────────────────

class AttendanceUpdateRequest(BaseModel):
    check_type:    Optional[str] = None   # "check_in" | "check_out"
    timestamp:     Optional[str] = None   # "YYYY-MM-DD HH:MM:SS" hoặc ISO
    note:          Optional[str] = None


class AttendanceCreateRequest(BaseModel):
    emp_code:   str
    check_type: str                       # "check_in" | "check_out"
    timestamp:  str                       # "YYYY-MM-DD HH:MM:SS" hoặc ISO
    note:       Optional[str] = ""


# ── GET /api/attendance/{log_id} ─────────────────────────────────

@router.get("/attendance/{log_id}")
def get_attendance_log(
    log_id: int,
    current_user=Depends(get_current_user),
):
    """Lấy thông tin 1 bản ghi điểm danh theo ID."""
    log = get_log_by_id(log_id)
    if not log:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi điểm danh")
    return log


# ── PUT /api/attendance/{log_id} ─────────────────────────────────

@router.put("/attendance/{log_id}")
def edit_attendance_log(
    log_id: int,
    body: AttendanceUpdateRequest,
    current_user=Depends(get_current_user),
):
    """
    Chỉnh sửa bản ghi điểm danh.
    Yêu cầu quyền manager hoặc admin.
    """
    from app.services.auth_service import require_manager
    # Kiểm tra quyền manager/admin
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(
            status_code=403,
            detail="Chỉ quản lý (manager/admin) mới được chỉnh sửa điểm danh",
        )

    if body.check_type and body.check_type not in ("check_in", "check_out"):
        raise HTTPException(status_code=422, detail="check_type phải là 'check_in' hoặc 'check_out'")

    try:
        updated = update_attendance_log(
            log_id       = log_id,
            check_type   = body.check_type,
            timestamp_str= body.timestamp,
            note         = body.note,
            updated_by   = current_user.full_name or current_user.email,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if updated is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi điểm danh")

    return {"success": True, "message": "Đã cập nhật bản ghi điểm danh", "log": updated}


# ── DELETE /api/attendance/{log_id} ──────────────────────────────

@router.delete("/attendance/{log_id}")
def remove_attendance_log(
    log_id: int,
    current_user=Depends(get_current_user),
):
    """
    Xoá bản ghi điểm danh.
    Yêu cầu quyền admin.
    """
    if current_user.role != "admin":
        raise HTTPException(
            status_code=403,
            detail="Chỉ admin mới được xoá bản ghi điểm danh",
        )

    try:
        deleted = delete_attendance_log(log_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not deleted:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi điểm danh")

    return {"success": True, "message": f"Đã xoá bản ghi #{log_id}"}


# ── POST /api/attendance/manual ──────────────────────────────────

@router.post("/attendance/manual")
def add_manual_attendance(
    body: AttendanceCreateRequest,
    current_user=Depends(get_current_user),
):
    """
    Tạo thủ công bản ghi điểm danh bù (quản lý thêm khi nhân viên quên chấm).
    Yêu cầu quyền manager hoặc admin.
    """
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(
            status_code=403,
            detail="Chỉ quản lý (manager/admin) mới được tạo điểm danh thủ công",
        )

    if body.check_type not in ("check_in", "check_out"):
        raise HTTPException(status_code=422, detail="check_type phải là 'check_in' hoặc 'check_out'")

    try:
        new_log = create_manual_attendance_log(
            emp_code    = body.emp_code,
            check_type  = body.check_type,
            timestamp_str = body.timestamp,
            note        = body.note or "",
            created_by  = current_user.full_name or current_user.email,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if new_log is None:
        raise HTTPException(
            status_code=404,
            detail=f"Không tìm thấy nhân viên với mã '{body.emp_code}' hoặc tài khoản đã bị vô hiệu hoá",
        )

    return {"success": True, "message": "Đã tạo bản ghi điểm danh thủ công", "log": new_log}