from fastapi import APIRouter
from fastapi.responses import FileResponse, JSONResponse
from datetime import datetime, timedelta
from app.database import SessionLocal, AttendanceLog, Employee
from app.attendance import get_logs_by_date, get_summary_today
import os

router = APIRouter(prefix="/api", tags=["reports"])


# ──────────────────────────────────────────
# GET /api/attendance  — Lịch sử chấm công
# ──────────────────────────────────────────
@router.get("/attendance")
def get_attendance(
    date:     str = None,       # YYYY-MM-DD, mặc định hôm nay
    emp_code: str = None,
    days:     int = 1           # Lấy N ngày gần nhất nếu không có date
):
    if date:
        logs = get_logs_by_date(date, emp_code)
    else:
        # Lấy N ngày gần nhất
        all_logs = []
        for i in range(days):
            d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            all_logs.extend(get_logs_by_date(d, emp_code))
        logs = all_logs
    return {"logs": logs, "total": len(logs)}


# ──────────────────────────────────────────
# GET /api/summary  — Thống kê hôm nay
# ──────────────────────────────────────────
@router.get("/summary")
def summary_today():
    return get_summary_today()


# ──────────────────────────────────────────
# GET /api/summary/range  — Thống kê theo khoảng thời gian
# ──────────────────────────────────────────
@router.get("/summary/range")
def summary_range(from_date: str, to_date: str):
    db = SessionLocal()
    try:
        start = datetime.strptime(from_date, "%Y-%m-%d").replace(hour=0, minute=0)
        end   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59)

        logs  = db.query(AttendanceLog).filter(
            AttendanceLog.timestamp >= start,
            AttendanceLog.timestamp <= end,
        ).all()

        # Group theo ngày
        by_date = {}
        for log in logs:
            day = log.timestamp.strftime("%Y-%m-%d")
            if day not in by_date:
                by_date[day] = {"check_in": set(), "check_out": set()}
            by_date[day][log.check_type].add(log.emp_code)

        result = []
        for day in sorted(by_date.keys()):
            result.append({
                "date":         day,
                "checked_in":   len(by_date[day]["check_in"]),
                "checked_out":  len(by_date[day]["check_out"]),
            })

        # Thống kê theo phòng ban
        dept_stats = {}
        for log in logs:
            dept = log.department or "Chưa xác định"
            dept_stats[dept] = dept_stats.get(dept, 0) + 1

        return {
            "from_date":   from_date,
            "to_date":     to_date,
            "total_logs":  len(logs),
            "by_date":     result,
            "by_dept":     [{"dept": k, "count": v} for k, v in dept_stats.items()],
        }
    finally:
        db.close()


# ──────────────────────────────────────────
# GET /api/reports/export  — Xuất Excel
# ──────────────────────────────────────────
@router.get("/reports/export")
def export_excel(from_date: str, to_date: str):
    """Xuất báo cáo Excel theo khoảng thời gian"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"error": "Cài openpyxl: pip install openpyxl"}, 500)

    db = SessionLocal()
    try:
        start = datetime.strptime(from_date, "%Y-%m-%d").replace(hour=0)
        end   = datetime.strptime(to_date,   "%Y-%m-%d").replace(hour=23, minute=59)
        logs  = db.query(AttendanceLog).filter(
            AttendanceLog.timestamp >= start,
            AttendanceLog.timestamp <= end,
        ).order_by(AttendanceLog.timestamp).all()
    finally:
        db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo chấm công"

    # Tiêu đề
    ws.merge_cells("A1:G1")
    ws["A1"] = f"BÁO CÁO CHẤM CÔNG  —  {from_date} đến {to_date}"
    ws["A1"].font      = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header
    headers = ["STT", "Mã NV", "Họ tên", "Phòng ban", "Loại", "Thời gian", "Trạng thái"]
    thin    = Side(border_style="thin", color="CCCCCC")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1A365D")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    # Dữ liệu
    for i, log in enumerate(logs, 1):
        row   = i + 2
        check = "Vào" if log.check_type == "check_in" else "Ra"
        vals  = [
            i,
            log.emp_code,
            log.emp_name,
            log.department,
            check,
            log.timestamp.strftime("%H:%M:%S  %d/%m/%Y"),
            log.note or "",
        ]
        fill_color = "F0FFF4" if log.check_type == "check_in" else "EBF4FF"
        row_fill   = PatternFill("solid", fgColor=fill_color)
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
            cell.fill      = row_fill

    # Độ rộng cột
    widths = [6, 10, 22, 18, 8, 24, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Lưu file
    os.makedirs("data/exports", exist_ok=True)
    filename = f"data/exports/chamcong_{from_date}_{to_date}.xlsx"
    wb.save(filename)

    return FileResponse(
        filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"BaoCaoChamCong_{from_date}_{to_date}.xlsx"
    )
